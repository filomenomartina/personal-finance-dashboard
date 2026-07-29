#!/usr/bin/env python3
"""
make_sample_workbook.py — generates `finance-tracker-SAMPLE.xlsx`, a fully
synthetic finance workbook in the exact schema the dashboard reads at runtime.
All figures are fake but internally consistent. Nothing here is real data.

    python3 make_sample_workbook.py            # -> finance-tracker-SAMPLE.xlsx

Sheets: Dashboard, Accounts, Historical, Expenses, Bills, Daily Log, Transactions.
The dashboard loads this file entirely in the browser; no data leaves your device.
"""
import random, datetime as dt
import openpyxl
from openpyxl.styles import Font

random.seed(42)
OUT = "finance-tracker-SAMPLE.xlsx"
BOLD = Font(bold=True)

CATS = {
    "Home":            ["Rent", "Groceries", "Utilities", "Internet/TV", "Miscellaneous"],
    "Transport":       ["Fuel", "Rail", "Taxi"],
    "Me":              ["Subscriptions", "Health", "Clothes", "Barber"],
    "Fun":             ["Travel", "Concerts", "Leisure"],
    "Food drinks out": ["Eating out", "Drinks", "Take away"],
    "Family":          ["Gifts", "Support"],
    "Work":            ["Equipment", "Fees"],
}
MERCHANTS = {
    "Rent": ["RIVERSIDE LETTINGS"], "Groceries": ["FRESHMART", "GREENGROCER CO", "COSTSAVE"],
    "Utilities": ["BRIGHT ENERGY", "AQUAFLOW WATER"], "Internet/TV": ["STREAMNET", "FIBREONE"],
    "Miscellaneous": ["HOMESTORE", "FIXIT HARDWARE"],
    "Fuel": ["FUELSTOP", "ECOCHARGE"], "Rail": ["RAILCONNECT", "METRO TRAVEL"], "Taxi": ["CITYCABS", "RIDENOW"],
    "Subscriptions": ["STREAMFLIX", "CLOUDTUNES", "APPSTORE CO"], "Health": ["WELLNESS CLINIC", "PHARMA PLUS"],
    "Clothes": ["URBAN THREADS", "SHOE DEPOT"], "Barber": ["KINGS BARBERS"],
    "Travel": ["SKYJET AIR", "CITY HOTELS"], "Concerts": ["TICKETLINE", "ARENA LIVE"], "Leisure": ["CINEPLEX", "BOULDER GYM"],
    "Eating out": ["THE OLD KITCHEN", "BISTRO 21", "NOODLE BAR"], "Drinks": ["THE ANCHOR PUB", "CRAFT TAPROOM"],
    "Take away": ["DELIVEROO CO", "PIZZA EXPRESS CO"],
    "Gifts": ["GIFTBOX CO"], "Support": ["A. RELATIVE"],
    "Equipment": ["TECHWORLD", "OFFICE DEPOT CO"], "Fees": ["PRO BODY FEES"],
}
TYP = {
    "Rent": 950, "Groceries": 380, "Utilities": 140, "Internet/TV": 70, "Miscellaneous": 90,
    "Fuel": 70, "Rail": 120, "Taxi": 45,
    "Subscriptions": 55, "Health": 90, "Clothes": 110, "Barber": 22,
    "Travel": 180, "Concerts": 90, "Leisure": 80,
    "Eating out": 220, "Drinks": 160, "Take away": 70,
    "Gifts": 40, "Support": 60, "Equipment": 50, "Fees": 30,
}
CONF = ["existing", "learned", "learned", "existing", "low"]

def month_range(a, b):
    y, m = a; ey, em = b; out = []
    while (y, m) <= (ey, em):
        out.append(f"{y:04d}-{m:02d}"); m += 1
        if m > 12: m = 1; y += 1
    return out
MONTHS = month_range((2023, 7), (2026, 6))

# ---- transactions ----
tx = []
ACCTS = ["Sample Bank", "Everyday Card", "Rewards Card"]
for ym in MONTHS:
    y, m = int(ym[:4]), int(ym[5:])
    for cat, subs in CATS.items():
        for sub in subs:
            if random.random() < 0.12: continue
            amt = round(TYP[sub] * random.uniform(0.6, 1.5), 2)
            day = random.randint(1, 28)
            tx.append([dt.datetime(y, m, day), ym, random.choice(ACCTS),
                       random.choice(MERCHANTS[sub]), amt, cat, sub, "Recurring", random.choice(CONF)])
    if random.random() < 0.3:  # occasional refund (negative)
        tx.append([dt.datetime(y, m, random.randint(1, 28)), ym, "Sample Bank", "REFUND CO",
                   -round(random.uniform(20, 120), 2), "Home", "Miscellaneous", "Recurring", "existing"])
# a cluster of one-off capital outlays
for ym, label, cat, sub, amt in [
    ("2026-01", "FURNISH CO", "Home", "Miscellaneous", 4200),
    ("2026-02", "INTERIORS LTD", "Home", "Miscellaneous", 2600),
    ("2026-03", "HMRC SAMPLE", "Work", "Fees", 1800)]:
    y, m = int(ym[:4]), int(ym[5:])
    tx.append([dt.datetime(y, m, 12), ym, "Sample Bank", label, amt, cat, sub, "One-off", "existing"])

# ---- monthly aggregates (for Historical) ----
cats = list(CATS.keys())
mrec = {ym: 0.0 for ym in MONTHS}
for t in tx:
    if t[7] == "Recurring": mrec[t[1]] += t[4]
hist = []
for i, ym in enumerate(MONTHS):
    pay = round(4200 + i * 22 + random.randint(-200, 500), 2)
    exp = round(mrec[ym], 2)
    netsav = round(max(0, pay - exp - random.uniform(200, 900)), 2)
    savpct = round(netsav / pay, 4) if pay else 0
    cashbal = round(15000 + i * 250 + random.randint(-800, 800), 2)
    hist.append((ym, cashbal, pay, exp, netsav, savpct))

wb = openpyxl.Workbook()

def title(ws, t): ws["A1"] = t; ws["A1"].font = BOLD

# Dashboard — label(A) / value(B)
ws = wb.active; ws.title = "Dashboard"; title(ws, "FINANCE DASHBOARD (SAMPLE)")
ws["A2"] = "Note"; ws["B2"] = "All figures are illustrative dummy data."
for i, (k, v) in enumerate([
    ("Net pay", 4800), ("Payday", 28), ("Month opening net liquid", 21000),
    ("Cash reserve goal", 20000), ("Cash reserve monthly target", 600),
    ("Cash savings rate", 0.18), ("All-in wealth rate", 0.31),
    ("Footer note", "Sample data — replace with your own workbook."),
]):
    ws.cell(4 + i, 1, k); ws.cell(4 + i, 2, v)

# Accounts — name(A) / value(D); section headers in A; SUMMARY block for totals
ws = wb.create_sheet("Accounts"); title(ws, "ACCOUNTS (SAMPLE)")
ws["A2"] = "Last updated"; ws["B2"] = dt.datetime(2026, 6, 30)
r = 4
def sec(ws, r, name, rows):
    ws.cell(r, 1, name).font = BOLD; r += 1
    for n, v in rows:
        ws.cell(r, 1, n); ws.cell(r, 4, v); r += 1
    return r + 1
r = sec(ws, r, "CASH", [("Sample Bank", 6200), ("Everyday Card", 0), ("Rewards Card", 120)])
r = sec(ws, r, "SAVINGS", [("Instant saver", 16000), ("Premium bonds", 5000)])
r = sec(ws, r, "REFUNDS DUE", [("Deposit refund", 300), ("Expense claim", 165)])
r = sec(ws, r, "INVESTMENTS", [("Stocks & shares ISA", 24000), ("Pension pot", 52000),
                               ("Index funds", 18000), ("Crypto", 3500), ("Home (property)", 420000), ("Company shareholding", 45000)])
r = sec(ws, r, "LIABILITIES", [("Credit card A", 3200), ("Credit card B", 2400),
                               ("Personal loan", 5000), ("Mortgage (outstanding)", 260000)])
# summary block (cur reset so these aren't added to a section, but findByLabel reads them)
ws.cell(r, 1, "SUMMARY").font = BOLD; r += 1
liquid = 6200 + 120 + 16000 + 5000 + 300 + 165
inv = 24000 + 52000 + 18000 + 3500 + 420000 + 45000
liab = 3200 + 2400 + 5000 + 260000
ws.cell(r, 1, "Company valuation"); ws.cell(r, 2, 500000); r += 1
for k, v in [("Total liquid", liquid), ("Total investments", inv),
             ("Total liabilities", liab), ("Gross assets", liquid + inv),
             ("Net worth", liquid + inv - liab)]:
    ws.cell(r, 1, k); ws.cell(r, 4, v); r += 1

# Historical — month(A) cash(B) pay(C) exp(D) netsav(E) savpct(F)
ws = wb.create_sheet("Historical"); title(ws, "HISTORICAL — monthly (SAMPLE)")
for j, h in enumerate(["Month", "Cash balance", "Net pay", "Expenses", "Net savings", "Savings %"], 1):
    ws.cell(3, j, h).font = BOLD
for i, row in enumerate(hist):
    for j, v in enumerate(row, 1): ws.cell(4 + i, j, v)

# Expenses — reporting month + Category/Budget block
ws = wb.create_sheet("Expenses"); title(ws, "EXPENSES — budget vs actual (SAMPLE)")
ws["A2"] = "Reporting month"; ws["B2"] = MONTHS[-1]
ws.cell(4, 1, "Category").font = BOLD; ws.cell(4, 2, "Budget").font = BOLD
last = MONTHS[-1]
actual_by_cat = {c: 0.0 for c in cats}
for t in tx:
    if t[1] == last and t[7] == "Recurring": actual_by_cat[t[5]] += t[4]
for i, c in enumerate(cats):
    ws.cell(5 + i, 1, c); ws.cell(5 + i, 2, round(max(50, actual_by_cat[c]) * random.uniform(0.9, 1.25)))
ws.cell(5 + len(cats), 1, "Total").font = BOLD

# Bills — header 'Bill'; name(A) amt(B) day(C) ... fixed flag(H)
ws = wb.create_sheet("Bills"); title(ws, "RECURRING BILLS (SAMPLE)")
ws.cell(3, 1, "Bill").font = BOLD
for j, h in enumerate(["Amount", "DD day", "To pay", "Next", "Days", "", "Fixed"], 2):
    ws.cell(3, j, h).font = BOLD
BILLS = [("Mortgage", 1180, 4, "F"), ("Council tax", 188, 1, "F"), ("Water", 34, 1, "F"),
         ("Gas & electric", 135, 1, "F"), ("Broadband & TV", 67, 17, "F"), ("Mobile", 30, 25, "F"),
         ("Life cover", 42, 1, "F"), ("Gym", 38, 9, ""), ("Cloud storage", 8, 29, ""),
         ("Music", 11, 21, ""), ("Personal loan", 96, 29, "F")]
for i, (n, amt, day, fx) in enumerate(BILLS):
    row = 4 + i
    ws.cell(row, 1, n); ws.cell(row, 2, amt); ws.cell(row, 3, day)
    ws.cell(row, 8, fx)  # col H fixed flag

# Daily Log — Date(A) + Net Liquid column; month-opening snapshots
ws = wb.create_sheet("Daily Log"); title(ws, "DAILY LOG (SAMPLE)")
for j, h in enumerate(["Date", "Net Liquid", "Investments", "Liabilities", "Net Worth", "Note"], 1):
    ws.cell(3, j, h).font = BOLD
DL = [(dt.datetime(2026, 4, 1), 19400), (dt.datetime(2026, 5, 1), 20250),
      (dt.datetime(2026, 6, 1), 29500), (dt.datetime(2026, 6, 30), 31200)]
for i, (d, nl) in enumerate(DL):
    ws.cell(4 + i, 1, d); ws.cell(4 + i, 2, nl); ws.cell(4 + i, 6, "Month opening (sample)")

# Transactions — the ledger
ws = wb.create_sheet("Transactions")
for j, h in enumerate(["Date", "Month", "Account", "Description", "Amount",
                       "List 1", "List 2", "Bucket", "Confidence"], 1):
    ws.cell(1, j, h).font = BOLD
tx.sort(key=lambda t: t[0])
for i, t in enumerate(tx):
    for j, v in enumerate(t, 1): ws.cell(2 + i, j, v)

# Config — optional. WEALTH GROUPS map investment lines to donut segments;
# CATEGORIES pin spending-category colours/order. Omit the whole sheet and the
# dashboard auto-derives both from your data.
ws = wb.create_sheet("Config"); title(ws, "CONFIG — read by the dashboard (optional; auto-derived if absent)")
ws.cell(3, 1, "WEALTH GROUPS").font = BOLD
for j, h in enumerate(["Match (investment name contains; comma-separated)", "Label", "Colour"], 1):
    ws.cell(4, j, h).font = BOLD
for i, row in enumerate([("company", "Company stake", "#a78bfa"),
                         ("pension", "Pension", "#f59e0b"),
                         ("home,property", "Home", "#5b9cf6")]):
    for j, v in enumerate(row, 1): ws.cell(5 + i, j, v)
ws.cell(9, 1, "CATEGORIES").font = BOLD
for j, h in enumerate(["Category", "Colour", "Order"], 1):
    ws.cell(10, j, h).font = BOLD
for i, (c, col) in enumerate([("Home", "#5b9cf6"), ("Transport", "#2dd4bf"), ("Me", "#4ade80"),
                              ("Fun", "#a78bfa"), ("Food drinks out", "#f472b6"),
                              ("Family", "#f87171"), ("Work", "#fbbf24")]):
    ws.cell(11 + i, 1, c); ws.cell(11 + i, 2, col); ws.cell(11 + i, 3, i + 1)
ws.column_dimensions["A"].width = 46

wb.save(OUT)
print(f"Wrote {OUT}: {len(tx)} transactions, {len(MONTHS)} months, {len(wb.sheetnames)} sheets")
print("Sheets:", wb.sheetnames)
