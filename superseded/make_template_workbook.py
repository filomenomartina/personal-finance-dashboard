#!/usr/bin/env python3
"""
make_template_workbook.py — writes `finance-tracker-TEMPLATE.xlsx`: an EMPTY
workbook in the exact schema the dashboard reads, with the required sheet names,
section headers and column headers in place and one clearly-marked example row
per section. Delete the example rows, keep the headers, and fill in your own.

See SCHEMA.md for the full contract.  Run:  python3 make_template_workbook.py
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill

BOLD = Font(bold=True)
GREY = Font(italic=True, color="888888")
NOTEFILL = PatternFill("solid", fgColor="FFF6D5")
OUT = "finance-tracker-TEMPLATE.xlsx"

wb = openpyxl.Workbook()

def note(ws, text):
    ws["A1"] = text; ws["A1"].font = BOLD; ws["A1"].fill = NOTEFILL

# Dashboard — optional label/value pairs (col A / col B)
ws = wb.active; ws.title = "Dashboard"
note(ws, "DASHBOARD — optional settings. Label in column A, value in column B.")
rows = [("Net pay", "e.g. 3500  (your monthly take-home)"),
        ("Payday", "e.g. 28  (day of month salary lands)"),
        ("Month opening net liquid", "e.g. 5000"),
        ("Cash reserve goal", "e.g. 15000"),
        ("Cash reserve monthly target", "e.g. 500"),
        ("Cash savings rate", "e.g. 0.15"),
        ("All-in wealth rate", "e.g. 0.25"),
        ("Footer note", "Any note to show in the dashboard footer")]
for i, (k, v) in enumerate(rows):
    ws.cell(3 + i, 1, k); c = ws.cell(3 + i, 2, v); c.font = GREY

# Accounts — REQUIRED. Section headers in col A; item name in col A, value in col D.
ws = wb.create_sheet("Accounts")
note(ws, "ACCOUNTS — REQUIRED. Section headers (CASH/SAVINGS/REFUNDS DUE/INVESTMENTS/LIABILITIES) in col A; item name in col A, value in col D.")
r = 3
def sec(r, name, examples):
    ws.cell(r, 1, name).font = BOLD; r += 1
    for nm, val in examples:
        ws.cell(r, 1, nm).font = GREY; ws.cell(r, 4, val).font = GREY; r += 1
    return r + 1
r = sec(r, "CASH", [("Example current account (replace)", 0)])
r = sec(r, "SAVINGS", [("Example savings (replace)", 0)])
r = sec(r, "REFUNDS DUE", [("Example refund owed to you (replace)", 0)])
r = sec(r, "INVESTMENTS", [("Example home/property (replace)", 0), ("Example pension (replace)", 0), ("Example other investment (replace)", 0)])
r = sec(r, "LIABILITIES", [("Example mortgage (replace)", 0), ("Example credit card (replace)", 0)])
ws.cell(r, 1, "SUMMARY").font = BOLD; r += 1
for k in ["Total liquid", "Total investments", "Total liabilities", "Gross assets", "Net worth"]:
    ws.cell(r, 1, k).font = GREY; ws.cell(r, 4, 0).font = GREY; r += 1

# Transactions — REQUIRED ledger.
ws = wb.create_sheet("Transactions")
for j, h in enumerate(["Date", "Month", "Account", "Description", "Amount",
                       "List 1", "List 2", "Bucket", "Confidence"], 1):
    ws.cell(1, j, h).font = BOLD
ex = [dt.datetime(2026, 1, 15), "2026-01", "Example account", "EXAMPLE SHOP (replace)",
      42.50, "Home", "Groceries", "Recurring", "existing"]
for j, v in enumerate(ex, 1):
    ws.cell(2, j, v).font = GREY
ws.cell(3, 4, "Amount POSITIVE = money out; negative = refund. Categories are yours (List 1 / List 2).").font = GREY

# Historical — one row per month.
ws = wb.create_sheet("Historical")
for j, h in enumerate(["Month", "Cash balance", "Net pay", "Expenses", "Net savings", "Savings %"], 1):
    ws.cell(3, j, h).font = BOLD
for j, v in enumerate(["2026-01", 0, 0, 0, 0, 0], 1):
    ws.cell(4, j, v).font = GREY

# Expenses — budget vs actual.
ws = wb.create_sheet("Expenses")
ws["A1"] = "Reporting month"; ws["B1"] = "2026-01"
ws.cell(3, 1, "Category").font = BOLD; ws.cell(3, 2, "Budget").font = BOLD
ws.cell(4, 1, "Home (replace)").font = GREY; ws.cell(4, 2, 0).font = GREY
ws.cell(5, 1, "Total").font = BOLD

# Bills — header 'Bill'; name A, amount B, DD day C, fixed flag H.
ws = wb.create_sheet("Bills")
ws.cell(3, 1, "Bill").font = BOLD
for j, h in enumerate(["Amount", "DD day", "", "", "", "", "Fixed"], 2):
    ws.cell(3, j, h).font = BOLD
ws.cell(4, 1, "Example bill (replace)").font = GREY
ws.cell(4, 2, 0).font = GREY; ws.cell(4, 3, 1).font = GREY; ws.cell(4, 8, "F").font = GREY

# Daily Log — Date + Net Liquid.
ws = wb.create_sheet("Daily Log")
for j, h in enumerate(["Date", "Net Liquid", "Investments", "Liabilities", "Net Worth", "Note"], 1):
    ws.cell(3, j, h).font = BOLD
ws.cell(4, 1, dt.datetime(2026, 1, 1)).font = GREY
ws.cell(4, 2, 0).font = GREY; ws.cell(4, 6, "Month opening (replace)").font = GREY

# Config — OPTIONAL. Delete this whole sheet and the dashboard auto-derives colours & groupings.
ws = wb.create_sheet("Config")
note(ws, "CONFIG — OPTIONAL. WEALTH GROUPS map investment lines to donut segments; CATEGORIES pin colours/order. Delete the sheet to auto-derive everything.")
ws.cell(3, 1, "WEALTH GROUPS").font = BOLD
for j, h in enumerate(["Match (investment name contains; comma-separated)", "Label", "Colour"], 1):
    ws.cell(4, j, h).font = BOLD
for i, row in enumerate([("home,property", "Home", "#5b9cf6"),
                         ("pension", "Pension", "#f59e0b"),
                         ("company,stake,equity", "Company / business", "#a78bfa")]):
    for j, v in enumerate(row, 1): ws.cell(5 + i, j, v).font = GREY
ws.cell(9, 1, "CATEGORIES").font = BOLD
for j, h in enumerate(["Category", "Colour", "Order"], 1):
    ws.cell(10, j, h).font = BOLD
for i, (c, col) in enumerate([("Example category (replace)", "#5b9cf6")]):
    ws.cell(11 + i, 1, c).font = GREY; ws.cell(11 + i, 2, col).font = GREY; ws.cell(11 + i, 3, 1).font = GREY
ws.column_dimensions["A"].width = 46

wb.save(OUT)
print(f"Wrote {OUT}: {len(wb.sheetnames)} sheets — {wb.sheetnames}")
