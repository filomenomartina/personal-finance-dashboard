#!/usr/bin/env python3
"""
make_workbooks.py
Generates the two workbooks that ship with this repo:

  finance-tracker-SAMPLE.xlsx    every sheet, filled with synthetic but
                                 internally consistent data — open this first to
                                 see a fully-populated dashboard.
  finance-tracker-TEMPLATE.xlsx  the same structure, blank: every sheet, section
                                 header and column header in place, with one
                                 clearly-marked example row per section. Delete
                                 the example rows, keep the headers, add yours.

Both carry a filled-in `Config` sheet, so they double as the reference for the
schema. Nothing here is real data.

Requires openpyxl.  pip install openpyxl

Run:  python3 make_workbooks.py            # both
      python3 make_workbooks.py sample
      python3 make_workbooks.py template
"""
import sys, random, datetime as dt
import openpyxl
from openpyxl.styles import Font

WHICH = sys.argv[1].lower() if len(sys.argv) > 1 else "both"
if WHICH not in ("both", "sample", "template"):
    sys.exit("usage: make_workbooks.py [both|sample|template]")

SAMPLE_OUT = "finance-tracker-SAMPLE.xlsx"
TEMPLATE_OUT = "finance-tracker-TEMPLATE.xlsx"

BOLD = Font(bold=True)
NOTE = Font(italic=True, color="808080")
TODAY = dt.date.today()

# --------------------------------------------------------------- taxonomy
CATS = {
    "Home":            ["Rent", "Groceries", "Utilities", "Internet/TV", "Maintenance"],
    "Transport":       ["Fuel", "Insurance", "Service", "Public transport"],
    "Me":              ["Subscriptions", "Health", "Clothes"],
    "Fun":             ["Travel", "Concerts", "Leisure"],
    "Food drinks out": ["Eating out", "Drinks", "Take away"],
    "Family":          ["Gifts", "Support"],
    "Work":            ["Equipment", "Fees"],
}
CATCOL = {
    "Home": "#5b9cf6", "Transport": "#2dd4bf", "Me": "#4ade80", "Fun": "#a78bfa",
    "Food drinks out": "#f472b6", "Family": "#f87171", "Work": "#fbbf24",
}
MERCHANTS = {
    "Rent": ["RIVERSIDE LETTINGS"],
    "Groceries": ["FRESHMART", "GREENGROCER CO", "COSTSAVE"],
    "Utilities": ["BRIGHT ENERGY", "AQUAFLOW WATER"],
    "Internet/TV": ["STREAMNET", "FIBREONE"],
    "Maintenance": ["FIXIT HARDWARE", "HOMESTORE"],
    "Fuel": ["FUELSTOP", "ECOCHARGE"],
    "Insurance": ["SAFEDRIVE INS"],
    "Service": ["AUTOCARE GARAGE"],
    "Public transport": ["METRO TRAVEL", "RAILCONNECT"],
    "Subscriptions": ["STREAMFLIX", "CLOUDTUNES"],
    "Health": ["WELLNESS CLINIC", "PHARMA PLUS"],
    "Clothes": ["URBAN THREADS", "SHOE DEPOT"],
    "Travel": ["SKYJET AIR", "CITY HOTELS"],
    "Concerts": ["TICKETLINE", "ARENA LIVE"],
    "Leisure": ["CINEPLEX", "BOULDER GYM"],
    "Eating out": ["THE OLD KITCHEN", "BISTRO 21"],
    "Drinks": ["THE ANCHOR PUB", "CRAFT TAPROOM"],
    "Take away": ["NOODLE BAR", "PIZZA CO"],
    "Gifts": ["GIFTBOX CO"],
    "Support": ["A. RELATIVE"],
    "Equipment": ["TECHWORLD", "OFFICE DEPOT CO"],
    "Fees": ["PRO BODY FEES"],
}
TYP = {
    "Rent": 950, "Groceries": 380, "Utilities": 140, "Internet/TV": 70, "Maintenance": 90,
    "Fuel": 70, "Insurance": 45, "Service": 60, "Public transport": 85,
    "Subscriptions": 55, "Health": 90, "Clothes": 110,
    "Travel": 180, "Concerts": 90, "Leisure": 80,
    "Eating out": 220, "Drinks": 160, "Take away": 70,
    "Gifts": 40, "Support": 60,
    "Equipment": 50, "Fees": 30,
}
CONF = ["existing", "learned", "learned", "existing", "low"]


def month_range(start, end):
    y, m = start
    ey, em = end
    out = []
    while (y, m) <= (ey, em):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


# 36 months ending with the current month, so there is always a "this month".
MONTHS = month_range((TODAY.year - 3, TODAY.month), (TODAY.year, TODAY.month))
CUR_M = MONTHS[-1]


def build_transactions():
    random.seed(42)
    tx = []
    accounts = ["Sample Bank", "Everyday Card", "Rewards Card"]
    for ym in MONTHS:
        y, m = int(ym[:4]), int(ym[5:])
        maxday = 28 if ym != CUR_M else max(1, min(28, TODAY.day))
        for cat, subs in CATS.items():
            for sub in subs:
                if random.random() < 0.15:
                    continue
                tx.append([dt.datetime(y, m, random.randint(1, maxday)), ym,
                           random.choice(accounts), random.choice(MERCHANTS[sub]),
                           round(TYP[sub] * random.uniform(0.6, 1.5), 2),
                           cat, sub, "Recurring", random.choice(CONF)])
        tx.append([dt.datetime(y, m, min(25, maxday)), ym, "Sample Bank",
                   "EMPLOYER PAYROLL", -round(4200 + random.uniform(-150, 400), 2),
                   "Income", "Salary", "Income", "existing"])
        tx.append([dt.datetime(y, m, min(26, maxday)), ym, "Sample Bank",
                   "TO SAVINGS", round(max(0, random.uniform(-200, 1100)), 2),
                   "Savings", "Transfer", "Savings", "existing"])
        if random.random() < 0.25:
            tx.append([dt.datetime(y, m, random.randint(1, maxday)), ym, "Sample Bank",
                       "REFUND CO", -round(random.uniform(20, 120), 2),
                       "Home", "Maintenance", "Recurring", "existing"])
        if random.random() < 0.3:
            tx.append([dt.datetime(y, m, random.randint(1, maxday)), ym, "Sample Bank",
                       "J. FRIEND", round(random.uniform(40, 300), 2),
                       "Me", "Health", "One-off", "review"])
    for off, label, cat, sub, amt in [
        (6, "FURNISH CO", "Home", "Maintenance", 12500),
        (5, "INTERIORS LTD", "Home", "Maintenance", 6400),
        (4, "TAX AUTHORITY", "Work", "Fees", 4200),
    ]:
        ym = MONTHS[max(0, len(MONTHS) - 1 - off)]
        y, m = int(ym[:4]), int(ym[5:])
        tx.append([dt.datetime(y, m, 12), ym, "Sample Bank", label, amt, cat, sub,
                   "One-off", "existing"])
    tx.sort(key=lambda t: t[0])
    return tx


def rollups(tx):
    R = {ym: {"rec": 0.0, "oneoff": 0.0, "income": 0.0, "tosav": 0.0,
              "cats": {c: 0.0 for c in CATS}} for ym in MONTHS}
    for t in tx:
        ym, amt, cat, bucket = t[1], t[4], t[5], t[7]
        r = R[ym]
        if bucket == "Recurring":
            r["rec"] += amt
            if cat in r["cats"]:
                r["cats"][cat] += amt
        elif bucket == "One-off":
            r["oneoff"] += amt
        elif bucket == "Income":
            r["income"] += -amt
        elif bucket == "Savings":
            r["tosav"] += amt
    return R


# Accounts: (section, [(name, value, due_in_days)]) — value goes in column D.
SAMPLE_ACCOUNTS = [
    ("CASH", [("Sample Bank", 6200, None), ("Everyday Card", 0, None),
              ("Rewards Card", 120, None)]),
    ("SAVINGS", [("Instant saver", 16000, None), ("Premium bonds", 5000, None)]),
    ("REFUNDS DUE", [("Deposit refund", 300, None), ("Expense claim", 165, None)]),
    ("INVESTMENTS", [("Stocks & shares ISA", 24000, None), ("Pension pot", 52000, None),
                     ("Index funds", 18000, None), ("Crypto", 3500, None),
                     ("Home (property)", 420000, None),
                     ("Company shareholding", 45000, None)]),
    ("LIABILITIES", [("Credit card A", 3200, 25), ("Credit card B", 2400, 40),
                     ("Personal loan", 5000, 200),
                     ("Mortgage (outstanding)", 260000, 365 * 18)]),
]
TEMPLATE_ACCOUNTS = [
    ("CASH", [("Example current account (replace)", 0, None)]),
    ("SAVINGS", [("Example savings (replace)", 0, None)]),
    ("REFUNDS DUE", [("Example refund owed to you (replace)", 0, None)]),
    ("INVESTMENTS", [("Example home/property (replace)", 0, None),
                     ("Example pension (replace)", 0, None),
                     ("Example other investment (replace)", 0, None)]),
    ("LIABILITIES", [("Example mortgage (replace)", 0, 365 * 18),
                     ("Example credit card (replace)", 0, 25)]),
]
SAMPLE_BILLS = [("Water", 34, 1, "F"), ("Council tax", 188, 1, "F"),
                ("Life cover", 82, 1, "F"), ("Gas & electric", 135, 1, "F"),
                ("Health plan", 123, 4, "F"), ("Mortgage", 1150, 4, "F"),
                ("Gym membership", 45, 9, ""), ("TV & broadband", 67, 17, "F"),
                ("Mobile", 30, 25, "F"), ("Cloud storage", 10, 29, ""),
                ("Personal loan", 73, 29, "F"), ("Credit card A", 250, 20, "")]
TEMPLATE_BILLS = [("Example bill — direct debit (replace)", 0, 1, "F"),
                  ("Example bill — variable (replace)", 0, 15, "")]


def build(blank):
    """blank=True -> the empty TEMPLATE; blank=False -> the filled SAMPLE."""
    tag = "TEMPLATE" if blank else "SAMPLE"
    tx = [] if blank else build_transactions()
    R = rollups(tx) if tx else {}
    accounts = TEMPLATE_ACCOUNTS if blank else SAMPLE_ACCOUNTS
    bills = TEMPLATE_BILLS if blank else SAMPLE_BILLS
    wb = openpyxl.Workbook()

    def note(ws, cell, text):
        ws[cell] = text
        ws[cell].font = NOTE

    # ---------------------------------------------------------- Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = f"DASHBOARD — OPTIONAL. Label in column A, value in column B. ({tag})"
    ws["A1"].font = BOLD
    # "Target card spend per month" is the forward parameter for the 12-month cash-flow
    # forecast: set it by hand and the forecast uses it verbatim. Leave the row out entirely
    # and the dashboard falls back to a run-rate computed from the Transactions ledger.
    rows = ([("Net pay", 0), ("Payday", 25), ("Cash reserve goal", 0),
             ("Cash reserve monthly target", 0), ("Month opening net liquid", 0),
             ("Target card spend per month", 0),
             ("All-in wealth rate", 0), ("Cash savings rate", 0),
             ("Footer note", "Replace with your own note, or delete this row.")]
            if blank else
            [("Net pay", 4200), ("Payday", 25), ("Cash reserve goal", 20000),
             ("Cash reserve monthly target", 800), ("Month opening net liquid", 26000),
             ("Target card spend per month", 1600),
             ("All-in wealth rate", 0.28), ("Cash savings rate", 0.14),
             ("Footer note", "All figures in this workbook are synthetic sample data.")])
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 32

    # ------------------------------------------------------------- Config
    ws = wb.create_sheet("Config")
    ws["A1"] = ("CONFIG — OPTIONAL. Delete the whole sheet and the dashboard derives "
                "everything from your data. Sections are matched by the headers below.")
    ws["A1"].font = BOLD
    r = 3
    ws.cell(r, 1, "SETTINGS").font = BOLD
    r += 1
    for k, v, hint in [
        ("Currency symbol", "£", "used everywhere, privacy mode included"),
        ("Locale", "en-GB", "date and number formatting, e.g. en-US, it-IT, de-DE"),
        ("Language", "", "UI language: en, it, es, fr or de. Blank = follow the browser"),
        ("Owner", "" if blank else "Sample User", "shown in the header; blank to hide"),
        ("History from", "", "YYYY-MM — trims the long-run charts; blank = all data"),
    ]:
        ws.cell(r, 1, k); ws.cell(r, 2, v); note(ws, f"D{r}", hint); r += 1

    r += 1
    ws.cell(r, 1, "WEALTH GROUPS").font = BOLD
    note(ws, f"D{r}", "groups INVESTMENTS lines into the wealth donut")
    r += 1
    for j, h in enumerate(["Match (investment name contains; comma-separated)",
                           "Label", "Colour"], 1):
        ws.cell(r, j, h).font = BOLD
    r += 1
    for match, label, colour in [("home,property", "Home", "#5b9cf6"),
                                 ("pension,sipp", "Pension", "#f59e0b"),
                                 ("company,stake,equity", "Company / business", "#a78bfa")]:
        ws.cell(r, 1, match); ws.cell(r, 2, label); ws.cell(r, 3, colour); r += 1

    r += 1
    ws.cell(r, 1, "CATEGORIES").font = BOLD
    note(ws, f"D{r}", "pins colour and display order; omitted categories still appear")
    r += 1
    for j, h in enumerate(["Category", "Colour", "Order"], 1):
        ws.cell(r, j, h).font = BOLD
    r += 1
    if blank:
        ws.cell(r, 1, "Example category (replace with a List 1 value)")
        ws.cell(r, 2, "#5b9cf6"); ws.cell(r, 3, 1); r += 1
    else:
        for i, c in enumerate(CATS, start=1):
            ws.cell(r, 1, c); ws.cell(r, 2, CATCOL[c]); ws.cell(r, 3, i); r += 1

    r += 1
    ws.cell(r, 1, "ACCOUNT GROUPS").font = BOLD
    note(ws, f"D{r}", "role 'card repayment' = a bill that clears a card, not a purchase")
    r += 1
    for j, h in enumerate(["Match (bill name contains; comma-separated)", "Role"], 1):
        ws.cell(r, j, h).font = BOLD
    r += 1
    ws.cell(r, 1, "card,credit"); ws.cell(r, 2, "card repayment"); r += 1

    r += 1
    ws.cell(r, 1, "THRESHOLDS").font = BOLD
    note(ws, f"D{r}", "tune when the dashboard warns you")
    r += 1
    for j, h in enumerate(["Name", "Value"], 1):
        ws.cell(r, j, h).font = BOLD
    r += 1
    for k, v, hint in [
        ("Runway warning months", 2, "fewer months of liquid runway than this -> amber"),
        ("Pace warning pct", 100, "spend vs day-proportional budget -> amber"),
        ("Pace alert pct", 110, "-> red"),
        ("Insight min spend", 40, "ignore categories smaller than this when detecting changes"),
        ("Insight min change pct", 15, "only report a change bigger than this"),
    ]:
        ws.cell(r, 1, k); ws.cell(r, 2, v); note(ws, f"D{r}", hint); r += 1
    for col, w in [("A", 46), ("B", 20), ("C", 12), ("D", 68)]:
        ws.column_dimensions[col].width = w

    # ----------------------------------------------------------- Accounts
    ws = wb.create_sheet("Accounts")
    ws["A1"] = (f"ACCOUNTS — REQUIRED. Section header in column A, then one row per "
                f"line item: name in A, value in D. ({tag})")
    ws["A1"].font = BOLD
    ws["A2"] = "Last updated"
    ws["B2"] = dt.datetime(TODAY.year, TODAY.month, TODAY.day)
    for j, h in enumerate(["Account", "Amount", "Currency", "GBP value",
                           "Due (liabilities)"], 1):
        ws.cell(4, j, h).font = BOLD
    note(ws, "G4", "column D is what the dashboard reads; column E splits short- from long-term debt")
    r = 6
    totals = {}
    for name, items in accounts:
        ws.cell(r, 1, name).font = BOLD
        r += 1
        s = 0
        for n, v, due in items:
            ws.cell(r, 1, n)
            ws.cell(r, 2, v)
            ws.cell(r, 3, "GBP")
            ws.cell(r, 4, v)
            if due is not None:
                d = TODAY + dt.timedelta(days=due)
                ws.cell(r, 5, dt.datetime(d.year, d.month, d.day))
            s += v
            r += 1
        totals[name] = s
        r += 1
    liquid = totals["CASH"] + totals["SAVINGS"] + totals["REFUNDS DUE"]
    invest, liab = totals["INVESTMENTS"], totals["LIABILITIES"]
    ws.cell(r, 1, "SUMMARY").font = BOLD
    note(ws, f"G{r}", "optional — recomputed from the line items when absent")
    r += 1
    ws.cell(r, 1, "Company valuation")
    ws.cell(r, 2, 0 if blank else 500000)
    r += 1
    for label, val in [("Total liquid", liquid), ("Total investments", invest),
                       ("Total liabilities", liab), ("Gross assets", liquid + invest),
                       ("Net worth", liquid + invest - liab)]:
        ws.cell(r, 1, label)
        ws.cell(r, 4, val)
        r += 1
    ws.column_dimensions["A"].width = 40

    # -------------------------------------------------------------- Bills
    ws = wb.create_sheet("Bills")
    ws["A1"] = f"BILLS — OPTIONAL. Recurring commitments. ({tag})"
    ws["A1"].font = BOLD
    for j, h in enumerate(["Bill", "Amount", "Direct debit day", "To pay (1/0)",
                           "Next collection", "Days until", "Notes", "Fixed?"], 1):
        ws.cell(3, j, h).font = BOLD
    note(ws, "J3", "column H: F/Y/1 marks a bill you cannot easily cancel")
    for i, (n, amt, day, fixed) in enumerate(bills):
        row = 4 + i
        nm = day - TODAY.day
        if nm <= 0:
            nm += 30
        coll = TODAY + dt.timedelta(days=nm)
        ws.cell(row, 1, n); ws.cell(row, 2, amt); ws.cell(row, 3, day)
        ws.cell(row, 4, 1)
        ws.cell(row, 5, dt.datetime(coll.year, coll.month, coll.day))
        ws.cell(row, 6, nm)
        ws.cell(row, 8, fixed)
    ws.column_dimensions["A"].width = 34

    # ----------------------------------------------------------- Expenses
    ws = wb.create_sheet("Expenses")
    ws["A1"] = f"EXPENSES — OPTIONAL. Monthly budget per category. ({tag})"
    ws["A1"].font = BOLD
    ws["A2"] = "Reporting month"
    ws["B2"] = CUR_M
    for j, h in enumerate(["Category", "Budget", "Actual", "Variance"], 1):
        ws.cell(4, j, h).font = BOLD
    note(ws, "F4", "actuals are recomputed from the ledger; the block ends at 'Total'")
    r = 5
    if blank:
        ws.cell(r, 1, "Example category (replace)"); ws.cell(r, 2, 0); r += 1
    else:
        for c in CATS:
            actual = round(R[CUR_M]["cats"][c], 2)
            ref = (sum(R[m]["cats"][c] for m in MONTHS[-13:-1]) / 12
                   if len(MONTHS) > 12 else max(actual, 50))
            budget = round(max(50, ref * random.uniform(0.95, 1.2)))
            ws.cell(r, 1, c); ws.cell(r, 2, budget); ws.cell(r, 3, actual)
            ws.cell(r, 4, round(budget - actual, 2))
            r += 1
    ws.cell(r, 1, "Total")
    ws.cell(r, 2, sum(ws.cell(x, 2).value or 0 for x in range(5, r)))
    ws.column_dimensions["A"].width = 32

    # --------------------------------------------------------- Historical
    ws = wb.create_sheet("Historical")
    ws["A1"] = f"HISTORICAL — OPTIONAL. One row per month, oldest first. ({tag})"
    ws["A1"].font = BOLD
    for j, h in enumerate(["Month", "Cash", "Pay", "Expenses", "Net savings",
                           "Savings rate", "Pension contrib"], 1):
        ws.cell(3, j, h).font = BOLD
    note(ws, "I3", "month as YYYY-MM text; savings rate as a fraction (0.25 = 25%)")
    if blank:
        ws.cell(4, 1, CUR_M)
        for j in range(2, 8):
            ws.cell(4, j, 0)
    else:
        cash = 12000.0
        for i, ym in enumerate(MONTHS):
            r0 = R[ym]
            exp = r0["rec"] + r0["oneoff"]
            cash += r0["income"] - exp - r0["tosav"]
            row = 4 + i
            ws.cell(row, 1, ym)
            ws.cell(row, 2, round(cash))
            ws.cell(row, 3, round(r0["income"]))
            ws.cell(row, 4, round(exp))
            ws.cell(row, 5, round(r0["tosav"]))
            if r0["income"]:
                ws.cell(row, 6, round(r0["tosav"] / r0["income"], 4))
            ws.cell(row, 7, 320)

    # ---------------------------------------------------------- Daily Log
    ws = wb.create_sheet("Daily Log")
    ws["A1"] = f"DAILY LOG — OPTIONAL. Dated snapshots; drives month-on-month deltas. ({tag})"
    ws["A1"].font = BOLD
    for j, h in enumerate(["Date", "Net Liquid", "Investments", "Liabilities",
                           "Net Worth", "Note"], 1):
        ws.cell(3, j, h).font = BOLD
    note(ws, "H3", "one row per month-end is plenty")
    netliq = liquid - (0 if blank else 3200 + 2400 + 5000)
    if blank:
        ws.cell(4, 1, dt.datetime(TODAY.year, TODAY.month, TODAY.day))
        for j in range(2, 6):
            ws.cell(4, j, 0)
        ws.cell(4, 6, "Example snapshot (replace)")
    else:
        d = TODAY.replace(day=1)
        snaps = []
        for _ in range(7):
            d = d - dt.timedelta(days=1)
            snaps.append(d)
            d = d.replace(day=1)
        snaps = list(reversed(snaps)) + [TODAY]
        for i, day in enumerate(snaps):
            row = 4 + i
            drift = (len(snaps) - i) * 900
            ws.cell(row, 1, dt.datetime(day.year, day.month, day.day))
            ws.cell(row, 2, round(netliq + drift))
            ws.cell(row, 3, invest)
            ws.cell(row, 4, liab)
            ws.cell(row, 5, round(liquid + invest - liab + drift))
            ws.cell(row, 6, "synthetic snapshot")

    # ---------------------------------------------------------- Wish List
    ws = wb.create_sheet("Wish List")
    ws["A1"] = f"WISH LIST — OPTIONAL. Affordability panel. ({tag})"
    ws["A1"].font = BOLD
    ws["A3"] = "Net liquidity"; ws["B3"] = 0 if blank else netliq
    ws["A4"] = "Provisions"; ws["B4"] = 0 if blank else 5200
    ws["A5"] = "Free cash"; ws["B5"] = 0 if blank else netliq - 5200
    ws["A6"] = "Affordability ratio"; ws["B6"] = 0.25
    note(ws, "D6", "share of free cash any single item may consume")
    ws["A8"] = "ID"; ws["A8"].font = BOLD
    for j, h in enumerate(["Item", "Category", "Cost"], 2):
        ws.cell(8, j, h).font = BOLD
    ws.cell(8, 20, "Verdict").font = BOLD
    wish = ([(1, "Example wish-list item (replace)", "Fun", 0, "")] if blank else
            [(1, "Road bike", "Fun", 1800, "OK"),
             (2, "Kitchen refit", "Home", 14000, "Hold — above single-item gate"),
             (3, "Laptop", "Work", 2200, "OK"),
             (4, "Holiday", "Fun", 3500, "Review")])
    for i, (wid, item, cat, cost, verdict) in enumerate(wish):
        row = 9 + i
        ws.cell(row, 1, wid); ws.cell(row, 2, item); ws.cell(row, 3, cat)
        ws.cell(row, 4, cost); ws.cell(row, 20, verdict)
    ws.column_dimensions["B"].width = 34

    # ------------------------------------------------------- Transactions
    ws = wb.create_sheet("Transactions")
    for j, h in enumerate(["Date", "Month", "Account", "Description", "Amount",
                           "List 1", "List 2", "Bucket", "Confidence"], 1):
        ws.cell(1, j, h).font = BOLD
    if blank:
        for j, v in enumerate([dt.datetime(TODAY.year, TODAY.month, 1), CUR_M,
                               "Example account", "EXAMPLE MERCHANT (replace)", 0,
                               "Example category", "Example subcategory",
                               "Recurring", "existing"], 1):
            ws.cell(2, j, v)
        note(ws, "K1", "Amount positive = money out. Bucket: Recurring / One-off / "
                       "Income / Savings / REVIEW")
    else:
        for i, t in enumerate(tx):
            for j, v in enumerate(t, 1):
                ws.cell(2 + i, j, v)
    ws.column_dimensions["D"].width = 26

    out = TEMPLATE_OUT if blank else SAMPLE_OUT
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  {len(tx)} transactions · sheets: {', '.join(wb.sheetnames)}")


if WHICH in ("both", "sample"):
    build(blank=False)
if WHICH in ("both", "template"):
    build(blank=True)
